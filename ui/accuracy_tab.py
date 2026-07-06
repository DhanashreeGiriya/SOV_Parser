"""
Auto-extracted module: ui/accuracy_tab.py
"""

from __future__ import annotations

import pandas as pd
import numpy as np
import streamlit as st
import openpyxl
import io

_NUMERIC_FIELDS = {
    "BuildingValue", "OtherValue", "ContentsValue", "TimeElementValue",
    "BIValue", "RiskCount", "NumUnits", "DaysCovered", "GrossArea",
    "NumberOfStories", "NumStories", "YearBuilt", "ConstructionCode",
    "OccupancyCode", "OccupancyType", "ClassCode", "Latitude", "Longitude",
    "BIPeriod", "RoofCoverYear", "Roof Year Built", "Sprinkler Availability",
}


_JOIN_CANDIDATES = [
    "LocationID", "LocNumber", "LocationName", "LocName",
    "Street", "StreetAddress", "PostalCode",
]


def _cell_similarity(val_clean, val_ref, field: str):
    """Return (score 0-100, match_type str)."""
    from fuzzywuzzy import fuzz as _fz

    def _blank(v):
        if v is None: return True
        try:
            if pd.isna(v): return True
        except Exception: pass
        return str(v).strip() in ("", "nan", "None", "NaN", "<NA>")

    cb, rb = _blank(val_clean), _blank(val_ref)
    if cb and rb: return 100.0, "both_null"
    if cb or rb:  return 0.0,   "one_null"

    sc = str(val_clean).strip()
    sr = str(val_ref).strip()
    if sc.lower() == sr.lower(): return 100.0, "exact"

    if field in _NUMERIC_FIELDS:
        try:
            nc = float(sc.replace(",", "").replace("$", ""))
            nr = float(sr.replace(",", "").replace("$", ""))
            if nr == 0 and nc == 0: return 100.0, "exact_zero"
            if nr == 0:             return 0.0,   "ref_zero"
            diff = abs(nc - nr) / abs(nr)
            return round(max(0.0, 100.0 - diff * 100.0), 1), "numeric"
        except ValueError:
            pass

    return float(_fz.token_sort_ratio(sc.lower(), sr.lower())), "fuzzy"


def _align_dataframes(cleaned_df: pd.DataFrame, ref_df: pd.DataFrame,
                      join_key: str | None):
    """Inner-join on join_key or fall back to row-position alignment."""
    if join_key and join_key in cleaned_df.columns and join_key in ref_df.columns:
        merged = pd.merge(
            cleaned_df.reset_index(drop=True),
            ref_df.reset_index(drop=True),
            on=join_key, how="inner", suffixes=("_clean", "_ref")
        )
        # Prefer _clean suffix for clean cols, _ref for ref cols
        a_c_cols, a_r_cols = {}, {}
        for f in cleaned_df.columns:
            if f in merged.columns:           a_c_cols[f] = merged[f]
            elif f"{f}_clean" in merged.columns: a_c_cols[f] = merged[f"{f}_clean"]
        for f in ref_df.columns:
            if f in merged.columns and f not in cleaned_df.columns:
                a_r_cols[f] = merged[f]
            elif f"{f}_ref" in merged.columns:
                a_r_cols[f] = merged[f"{f}_ref"]
            elif f in merged.columns:
                a_r_cols[f] = merged[f]
        return (pd.DataFrame(a_c_cols).reset_index(drop=True),
                pd.DataFrame(a_r_cols).reset_index(drop=True))

    n = min(len(cleaned_df), len(ref_df))
    return cleaned_df.iloc[:n].reset_index(drop=True), ref_df.iloc[:n].reset_index(drop=True)


def _build_field_data(a_clean, a_ref, mapped_fields):
    """
    Compute per-row (score, match_type, clean_val, ref_val) for every mapped field.
    Returns dict: field -> list of (score, type, clean_str, ref_str).
    """
    data = {}
    n = len(a_clean)
    for field, ref_col in mapped_fields.items():
        rows = []
        for i in range(n):
            vc = a_clean[field].iloc[i] if field in a_clean.columns else None
            vr = a_ref[ref_col].iloc[i] if ref_col in a_ref.columns else None
            s, t = _cell_similarity(vc, vr, field)
            rows.append((s, t,
                         "" if _is_blank_v(vc) else str(vc).strip(),
                         "" if _is_blank_v(vr) else str(vr).strip()))
        data[field] = rows
    return data


def _is_blank_v(v):
    if v is None: return True
    try:
        if pd.isna(v): return True
    except Exception: pass
    return str(v).strip() in ("", "nan", "None", "NaN", "<NA>")


def build_accuracy_excel(a_clean: pd.DataFrame, a_ref: pd.DataFrame,
                         mapped_fields: dict, field_data: dict) -> bytes:
    """
    3-sheet accuracy workbook:
      Sheet 1 — Column_Score_Summary
      Sheet 2 — Row_Level_Comparison  (CLEAN | REF | MATCH% triplets)
      Sheet 3 — Discrepancy_Detail    (all cells with score < 90%)
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import numpy as np

    n = len(a_clean)
    fields = list(mapped_fields.keys())

    # ── Style helpers ─────────────────────────────────────────────────────────
    BLUE_HDR  = PatternFill("solid", fgColor="1B3A6B")
    GRN_HDR   = PatternFill("solid", fgColor="065F46")
    AMB_HDR   = PatternFill("solid", fgColor="92400E")
    GREY_HDR  = PatternFill("solid", fgColor="374151")
    F_HDR     = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
    F_BOLD    = Font(bold=True, size=9, name="Calibri")
    F_REG     = Font(size=9, name="Calibri")
    GRN_FILL  = PatternFill("solid", fgColor="C6EFCE")
    YLW_FILL  = PatternFill("solid", fgColor="FFEB9C")
    RED_FILL  = PatternFill("solid", fgColor="FFC7CE")
    GRY_FILL  = PatternFill("solid", fgColor="F2F2F2")
    ALT_FILL  = PatternFill("solid", fgColor="F0F4F8")
    BLU_FILL  = PatternFill("solid", fgColor="DEEAF1")
    thin = Side(style="thin", color="CCCCCC")
    BDR  = Border(top=thin, bottom=thin, left=thin, right=thin)

    def sfill(s):
        if s is None: return GRY_FILL
        if s >= 90:   return GRN_FILL
        if s >= 70:   return YLW_FILL
        return RED_FILL

    def hc(ws, r, c, v, fill, font=None, center=True):
        cell = ws.cell(r, c, v)
        cell.fill = fill; cell.font = font or F_HDR; cell.border = BDR
        cell.alignment = Alignment(horizontal="center" if center else "left",
                                   vertical="center", wrap_text=True)
        return cell

    def dc(ws, r, c, v, fill=None, font=None, align="left", fmt=None):
        cell = ws.cell(r, c, v)
        if fill: cell.fill = fill
        cell.font = font or F_REG; cell.border = BDR
        cell.alignment = Alignment(horizontal=align, vertical="center")
        if fmt: cell.number_format = fmt
        return cell

    def set_title(ws, text, cols, fill_hex="1B3A6B"):
        ws.row_dimensions[1].height = 30
        ws.merge_cells(f"A1:{get_column_letter(cols)}1")
        c = ws["A1"]
        c.value = text
        c.fill = PatternFill("solid", fgColor=fill_hex)
        c.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")

    def set_subtitle(ws, text, cols, fill_hex="2E75B6"):
        ws.row_dimensions[2].height = 22
        ws.merge_cells(f"A2:{get_column_letter(cols)}2")
        c = ws["A2"]
        c.value = text
        c.fill = PatternFill("solid", fgColor=fill_hex)
        c.font = Font(color="FFFFFF", size=9, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")

    def autowidth(ws, max_w=40):
        for col_cells in ws.columns:
            w = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[get_column_letter(
                col_cells[0].column)].width = min(w + 2, max_w)

    # ── Summary stats per field ───────────────────────────────────────────────
    summary = []
    for f in fields:
        rows = field_data[f]
        scores = [r[0] for r in rows]
        avg = round(float(np.mean(scores)), 1)
        exact      = sum(1 for r in rows if r[1] == "exact")
        both_null  = sum(1 for r in rows if r[1] == "both_null")
        one_null   = sum(1 for r in rows if r[1] == "one_null")
        numeric_p  = sum(1 for r in rows if r[1] == "numeric")
        fuzzy_p    = sum(1 for r in rows if r[1] == "fuzzy")
        band = ("Excellent ≥90%" if avg >= 90 else
                "Good 70–89%"    if avg >= 70 else "Needs Review <70%")
        summary.append(dict(field=f, ref=mapped_fields[f], avg=avg, exact=exact,
                            both_null=both_null, one_null=one_null,
                            numeric=numeric_p, fuzzy=fuzzy_p, band=band))
    summary.sort(key=lambda x: x["avg"])
    overall = round(float(np.mean([s["avg"] for s in summary])), 1)
    n_exc  = sum(1 for s in summary if s["avg"] >= 90)
    n_rev  = sum(1 for s in summary if s["avg"] <  70)

    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 1 — Column Score Summary
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Column_Score_Summary"
    N_COLS_1 = 10
    set_title(ws1, "SOV Accuracy Report — Cleaned vs Reference  |  Column-Level Similarity Scores", N_COLS_1)
    set_subtitle(ws1,
        f"{n} rows  ·  {len(fields)} fields compared  ·  "
        f"Overall Score: {overall}%  ·  "
        f"Excellent (≥90%): {n_exc}   Needs Review (<70%): {n_rev}", N_COLS_1)

    col_hdrs = ["#", "Output Field", "Ref Column", "Score", "Band",
                "Exact Match", "Both Null", "One-Null Miss", "Numeric Prox", "Fuzzy Text"]
    col_ws   = [4, 22, 22, 9, 18, 12, 10, 14, 13, 10]
    for ci, (h, w) in enumerate(zip(col_hdrs, col_ws), 1):
        hc(ws1, 3, ci, h, BLUE_HDR)
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[3].height = 24

    for ri, s in enumerate(summary, 1):
        rn = ri + 3
        alt = ALT_FILL if ri % 2 == 0 else None
        dc(ws1, rn, 1, ri,       alt, align="center")
        dc(ws1, rn, 2, s["field"], alt, F_BOLD)
        dc(ws1, rn, 3, s["ref"],  alt, font=Font(size=9, name="Calibri", color="2E75B6"))
        # Score cell — coloured
        sc = ws1.cell(rn, 4, s["avg"])
        sc.fill = sfill(s["avg"]); sc.font = F_BOLD
        sc.alignment = Alignment(horizontal="center", vertical="center")
        sc.border = BDR; sc.number_format = '0.0"%"'
        # Band cell
        bc = ws1.cell(rn, 5, s["band"])
        bc.fill = sfill(s["avg"]); bc.font = F_BOLD
        bc.alignment = Alignment(horizontal="center", vertical="center"); bc.border = BDR
        for ci, v in enumerate([s["exact"], s["both_null"], s["one_null"],
                                  s["numeric"], s["fuzzy"]], 6):
            dc(ws1, rn, ci, v, alt, align="center")

    # Overall row
    lr = len(summary) + 4
    ws1.merge_cells(f"A{lr}:C{lr}")
    oc = ws1.cell(lr, 1, "OVERALL AVERAGE")
    oc.fill = BLUE_HDR; oc.font = F_HDR
    oc.alignment = Alignment(horizontal="center", vertical="center"); oc.border = BDR
    ws1.cell(lr, 2).border = BDR; ws1.cell(lr, 3).border = BDR
    oc2 = ws1.cell(lr, 4, overall)
    oc2.fill = sfill(overall); oc2.font = Font(bold=True, size=11, name="Calibri")
    oc2.alignment = Alignment(horizontal="center", vertical="center")
    oc2.border = BDR; oc2.number_format = '0.0"%"'
    for ci in range(5, N_COLS_1 + 1):
        ws1.cell(lr, ci).fill = BLU_FILL; ws1.cell(lr, ci).border = BDR
    ws1.freeze_panes = "A4"

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 2 — Row Level Comparison
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Row_Level_Comparison")
    N_COLS_2 = 1 + len(fields) * 3
    set_title(ws2,
        "Row-Level Comparison  |  Each field: CLEAN (transformed) → REF (expected) → MATCH%",
        N_COLS_2)
    set_subtitle(ws2,
        "Green ≥90%  ·  Yellow 70–89%  ·  Red <70%  ·  Score cells show numeric proximity for value fields, fuzzy similarity for text fields",
        N_COLS_2)

    # Key column
    hc(ws2, 3, 1, "LocationID\n(Key)", GREY_HDR)
    ws2.column_dimensions["A"].width = 13

    fld_col = {}   # field -> (clean_col_idx, ref_col_idx, score_col_idx)
    ci = 2
    for f in fields:
        hc(ws2, 3, ci,   f"CLEAN\n{f}",          BLUE_HDR)
        hc(ws2, 3, ci+1, f"REF\n{mapped_fields[f]}", GRN_HDR)
        hc(ws2, 3, ci+2, f"MATCH%\n{f}",          AMB_HDR)
        fld_col[f] = (ci, ci+1, ci+2)
        ws2.column_dimensions[get_column_letter(ci)].width   = 20
        ws2.column_dimensions[get_column_letter(ci+1)].width = 20
        ws2.column_dimensions[get_column_letter(ci+2)].width = 10
        ci += 3
    ws2.row_dimensions[3].height = 28

    # Determine LocationID column for the key
    loc_col = "LocationID" if "LocationID" in a_clean.columns else (
              "LocNumber"  if "LocNumber"  in a_clean.columns else None)

    for ri in range(n):
        rn = ri + 4
        alt = ALT_FILL if ri % 2 == 0 else None
        loc_val = str(a_clean[loc_col].iloc[ri]).strip() if loc_col else str(ri + 1)
        dc(ws2, rn, 1, loc_val, alt, align="center")
        for f in fields:
            sc_c, rc_c, mc_c = fld_col[f]
            score, stype, vc_str, vr_str = field_data[f][ri]
            dc(ws2, rn, sc_c, vc_str, alt)
            dc(ws2, rn, rc_c, vr_str, alt)
            mc = ws2.cell(rn, mc_c, round(score, 1))
            mc.fill = sfill(score)
            mc.font = Font(bold=(score < 70), size=9, name="Calibri")
            mc.alignment = Alignment(horizontal="center", vertical="center")
            mc.border = BDR; mc.number_format = '0.0"%"'

    # Average footer row
    avg_r = n + 4
    ac = ws2.cell(avg_r, 1, "AVG")
    ac.fill = BLUE_HDR; ac.font = F_HDR
    ac.alignment = Alignment(horizontal="center", vertical="center"); ac.border = BDR
    for f in fields:
        sc_c, rc_c, mc_c = fld_col[f]
        avg_f = round(float(np.mean([field_data[f][i][0] for i in range(n)])), 1)
        for cc in [sc_c, rc_c]:
            ws2.cell(avg_r, cc).fill = BLU_FILL; ws2.cell(avg_r, cc).border = BDR
        avc = ws2.cell(avg_r, mc_c, avg_f)
        avc.fill = sfill(avg_f); avc.font = F_BOLD
        avc.alignment = Alignment(horizontal="center", vertical="center")
        avc.border = BDR; avc.number_format = '0.0"%"'
    ws2.freeze_panes = "B4"

    # ════════════════════════════════════════════════════════════════════════
    # Sheet 3 — Discrepancy Detail
    # ════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Discrepancy_Detail")
    disc_hdrs = ["LocationID", "Field", "Cleaned Value", "Reference Value",
                 "Match%", "Band", "Match Method"]
    disc_w    = [12, 22, 30, 30, 10, 16, 15]
    set_title(ws3, "Discrepancy Detail — All row × field combinations with Match% below 90%", 7, "991B1B")
    set_subtitle(ws3,
        "Sorted by field then row. Use this sheet to identify systematic transformation errors.",
        7, "B91C1C")
    for ci, (h, w) in enumerate(zip(disc_hdrs, disc_w), 1):
        hc(ws3, 3, ci, h, PatternFill("solid", fgColor="374151"))
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.row_dimensions[3].height = 22

    dr = 4
    for f in fields:
        for i in range(n):
            score, stype, vc_str, vr_str = field_data[f][i]
            if score < 90:
                loc_val = str(a_clean[loc_col].iloc[i]).strip() if loc_col else str(i + 1)
                alt = ALT_FILL if dr % 2 == 0 else None
                dc(ws3, dr, 1, loc_val,  alt, align="center")
                dc(ws3, dr, 2, f,        alt, F_BOLD)
                dc(ws3, dr, 3, vc_str,   alt)
                dc(ws3, dr, 4, vr_str,   alt)
                ms = ws3.cell(dr, 5, round(score, 1))
                ms.fill = sfill(score); ms.font = F_BOLD
                ms.alignment = Alignment(horizontal="center", vertical="center")
                ms.border = BDR; ms.number_format = '0.0"%"'
                band = "Good" if score >= 70 else "Needs Review"
                dc(ws3, dr, 6, band, sfill(score), align="center")
                dc(ws3, dr, 7, stype, alt, align="center")
                dr += 1
    ws3.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_accuracy_tab():
    """Accuracy QA — compare cleaned output column-for-column against a reference file."""

    st.markdown(
        '<p style="font-size:.85rem;color:#1a1a2e;margin-bottom:.8rem">'
        'Upload a reference file that uses the same field names as the cleaned output. '
        'Columns are matched automatically by name. Scores use numeric proximity for value '
        'fields and fuzzy string similarity for text fields.</p>',
        unsafe_allow_html=True)

    if "phase3_result" not in st.session_state:
        st.info("Complete the **Transform** step first to generate a cleaned SOV.", icon="ℹ️")
        return

    cleaned_df = st.session_state["phase3_result"]["cleaned_df"]

    # ── Step 1: Upload reference ─────────────────────────────────────────────
    st.markdown("#### Step 1 — Upload Reference File")
    ref_file = st.file_uploader(
        "Reference file (.xlsx, .xls, or .csv) — must use AIR/RMS field names as column headers",
        type=["xlsx", "xls", "csv"], key="ref_upload")

    if ref_file is None:
        st.caption("Waiting for reference file…")
        return

    try:
        if ref_file.name.endswith(".csv"):
            ref_df = pd.read_csv(ref_file, dtype=str)
        else:
            ref_df = pd.read_excel(ref_file, dtype=str)
        ref_df.columns = [str(c).strip() for c in ref_df.columns]
    except Exception as e:
        st.error(f"Could not read reference file: {e}"); return

    # ── Auto-map: extract CLEAN fields and match to ref columns ─────────────
    # If cleaned_df already has plain field names (from session state),
    # match directly. Extract only cols that exist in both.
    matched   = {f: f for f in cleaned_df.columns if f in ref_df.columns}
    unmatched_clean = [f for f in cleaned_df.columns if f not in ref_df.columns]
    unmatched_ref   = [f for f in ref_df.columns   if f not in cleaned_df.columns]

    c1, c2, c3 = st.columns(3)
    c1.success(f"✅  {len(matched)} fields auto-matched by name")
    c2.info(f"ℹ️  {len(unmatched_clean)} cleaned fields not in reference")
    c3.warning(f"⚠️  {len(unmatched_ref)} reference fields not in cleaned output")

    # ── Show mapping table ───────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("#### Step 2 — Field Mapping")

    with st.expander(f"View auto-mapped fields ({len(matched)})", expanded=False):
        m_rows = [{"Cleaned Field": k, "Reference Column": v, "Status": "✅ Auto-matched"}
                  for k, v in matched.items()]
        if unmatched_clean:
            for f in unmatched_clean:
                m_rows.append({"Cleaned Field": f, "Reference Column": "—", "Status": "⚪ Not in reference"})
        if m_rows:
            st.dataframe(pd.DataFrame(m_rows), use_container_width=True,
                         height=min(len(m_rows) * 35 + 60, 380))

    if not matched:
        st.error("No matching columns found. Ensure the reference file uses AIR/RMS schema field names.", icon="🚫")
        return

    # ── Step 3: Row alignment ────────────────────────────────────────────────
    st.markdown("#### Step 3 — Row Alignment")
    auto_key = next((k for k in _JOIN_CANDIDATES
                     if k in cleaned_df.columns and k in ref_df.columns), None)
    key_options = ["Row position (no join key)"] + [
        c for c in cleaned_df.columns if c in ref_df.columns]
    def_idx = key_options.index(auto_key) if auto_key in key_options else 0
    key_choice = st.selectbox("Align rows by", key_options, index=def_idx,
                              help="Key column for inner-join alignment. Rows without a match are excluded.")
    join_key = None if key_choice == "Row position (no join key)" else key_choice

    a_clean, a_ref = _align_dataframes(cleaned_df, ref_df, join_key)
    n_rows = len(a_clean)

    if join_key:
        st.caption(f"Joined on **{join_key}** — {n_rows:,} matching rows found.")
    else:
        st.caption(f"Row-position alignment — {n_rows:,} rows compared.")

    # ── Step 4: Run scoring ──────────────────────────────────────────────────
    st.markdown("---")
    run_col, _ = st.columns([1, 2])
    run_btn = run_col.button("📊  Compute Accuracy Scores", use_container_width=True)

    if run_btn:
        with st.spinner(f"Scoring {n_rows} rows × {len(matched)} fields…"):
            field_data = _build_field_data(a_clean, a_ref, matched)
        st.session_state["acc_field_data"] = field_data
        st.session_state["acc_matched"]    = matched
        st.session_state["acc_a_clean"]    = a_clean
        st.session_state["acc_a_ref"]      = a_ref
        st.session_state["acc_n_rows"]     = n_rows
        st.session_state["acc_ref_name"]   = ref_file.name

    if "acc_field_data" not in st.session_state:
        return

    # Restore from session
    field_data = st.session_state["acc_field_data"]
    matched    = st.session_state["acc_matched"]
    a_clean    = st.session_state["acc_a_clean"]
    a_ref      = st.session_state["acc_a_ref"]
    n_rows     = st.session_state["acc_n_rows"]
    import numpy as _np

    # ── Summary metrics ──────────────────────────────────────────────────────
    field_scores = {f: round(float(_np.mean([r[0] for r in rows])), 1)
                    for f, rows in field_data.items()}
    overall  = round(float(_np.mean(list(field_scores.values()))), 1)
    n_exc    = sum(1 for s in field_scores.values() if s >= 90)
    n_good   = sum(1 for s in field_scores.values() if 70 <= s < 90)
    n_rev    = sum(1 for s in field_scores.values() if s < 70)

    def sc_col(s):
        if s >= 90: return "#10b981"
        if s >= 70: return "#f59e0b"
        return "#ef4444"

    st.markdown("---")
    st.markdown("#### Results")
    m1, m2, m3, m4, m5 = st.columns(5)
    m1.markdown(f"""
<div style="background:#ffffff;border:1px solid #dee2e6;border-radius:6px;
            padding:.9rem 1.1rem;text-align:center">
  <div style="font-size:.68rem;color:#6c757d;text-transform:uppercase;letter-spacing:.06em;margin-bottom:.3rem">Overall Score</div>
  <div style="font-size:2.2rem;font-weight:700;font-family:'IBM Plex Mono',monospace;
              color:{sc_col(overall)}">{overall}%</div>
</div>""", unsafe_allow_html=True)
    m2.metric("Rows Compared",      f"{n_rows:,}")
    m3.metric("Fields Compared",    f"{len(field_scores)}")
    m4.metric("Excellent (≥ 90%)",  f"{n_exc}")
    m5.metric("Needs Review (< 70%)", f"{n_rev}")

    # ── Per-column bar chart ─────────────────────────────────────────────────
    st.markdown("---")
    st.markdown("**Per-Field Accuracy** — sorted worst → best")
    bar_rows = []
    for f, score in sorted(field_scores.items(), key=lambda x: x[1]):
        rows = field_data[f]
        exact    = sum(1 for r in rows if r[1] == "exact")
        one_null = sum(1 for r in rows if r[1] == "one_null")
        band = "Excellent" if score >= 90 else ("Good" if score >= 70 else "Review needed")
        band_bg = {"Excellent":"#dcfce7","Good":"#fef3c7","Review needed":"#fee2e2"}[band]
        band_fg = {"Excellent":"#065f46","Good":"#92400e","Review needed":"#991b1b"}[band]
        ref_col = matched.get(f, "—")
        bar_rows.append(f"""
<div style="display:grid;grid-template-columns:190px 180px 1fr 70px 110px 80px;
            border-bottom:1px solid #e5e7eb;align-items:center;padding:.32rem 0">
  <div style="font-family:'IBM Plex Mono',monospace;font-size:.78rem;color:#1a1a2e;
              padding:0 .6rem;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">{f}</div>
  <div style="font-size:.72rem;color:#6b7280;padding:0 .4rem;overflow:hidden;
              text-overflow:ellipsis;white-space:nowrap">{ref_col}</div>
  <div style="padding:0 .6rem">
    <div style="background:#f3f4f6;border-radius:3px;height:7px">
      <div style="background:{sc_col(score)};width:{score}%;height:7px;border-radius:3px"></div>
    </div>
  </div>
  <div style="font-family:'IBM Plex Mono',monospace;font-size:.84rem;font-weight:700;
              color:{sc_col(score)};padding:0 .3rem;text-align:right">{score}%</div>
  <div style="padding:0 .4rem;text-align:center">
    <span style="background:{band_bg};color:{band_fg};font-size:.65rem;font-weight:600;
                 padding:.12rem .4rem;border-radius:3px">{band}</span>
  </div>
  <div style="font-size:.68rem;color:#9ca3af;padding:0 .3rem;text-align:center">
    {exact} exact
  </div>
</div>""")

    st.markdown(f"""
<div style="border:1px solid #e5e7eb;border-radius:6px;overflow:hidden">
  <div style="display:grid;grid-template-columns:190px 180px 1fr 70px 110px 80px;
              background:#f8fafc;border-bottom:2px solid #e5e7eb;padding:.4rem 0">
    <div style="font-size:.68rem;font-weight:700;color:#6b7280;text-transform:uppercase;
                letter-spacing:.06em;padding:0 .6rem">Field</div>
    <div style="font-size:.68rem;font-weight:700;color:#6b7280;text-transform:uppercase;
                letter-spacing:.06em;padding:0 .4rem">Ref Column</div>
    <div style="font-size:.68rem;font-weight:700;color:#6b7280;text-transform:uppercase;
                letter-spacing:.06em;padding:0 .6rem">Match Bar</div>
    <div style="font-size:.68rem;font-weight:700;color:#6b7280;text-transform:uppercase;
                letter-spacing:.06em;padding:0 .3rem;text-align:right">Score</div>
    <div style="font-size:.68rem;font-weight:700;color:#6b7280;text-transform:uppercase;
                letter-spacing:.06em;padding:0 .4rem;text-align:center">Band</div>
    <div style="font-size:.68rem;font-weight:700;color:#6b7280;text-transform:uppercase;
                letter-spacing:.06em;padding:0 .3rem;text-align:center">Exact</div>
  </div>
  {"".join(bar_rows)}
</div>""", unsafe_allow_html=True)

    # ── Discrepancy drill-down ───────────────────────────────────────────────
    st.markdown("---")
    st.markdown("**Discrepancy Detail** — rows with score < 90 %")
    disc_filter = st.selectbox("Filter by field", ["All fields"] + sorted(field_scores.keys()),
                               key="acc_disc_filter")
    loc_col = next((c for c in ["LocationID","LocNumber"] if c in a_clean.columns), None)
    disc_rows = []
    for f in (matched if disc_filter == "All fields" else {disc_filter: matched[disc_filter]}):
        for i, (score, stype, vc, vr) in enumerate(field_data[f]):
            if score < 90:
                loc_val = str(a_clean[loc_col].iloc[i]).strip() if loc_col else str(i+1)
                disc_rows.append({
                    "LocationID":     loc_val,
                    "Field":          f,
                    "Cleaned Value":  vc,
                    "Reference Value":vr,
                    "Match %":        round(score, 1),
                    "Method":         stype,
                })
    if disc_rows:
        disc_df = pd.DataFrame(disc_rows).sort_values(["Field","Match %"])

        def sty_score(v):
            try:
                f = float(v)
                if f >= 90: return "color:#065f46;font-weight:bold"
                if f >= 70: return "color:#92400e"
                return "color:#991b1b;font-weight:bold"
            except: return ""

        st.dataframe(disc_df.style.map(sty_score, subset=["Match %"]),
                     use_container_width=True,
                     height=min(len(disc_rows) * 35 + 60, 440))
        st.caption(f"{len(disc_rows):,} discrepancies across {disc_filter.lower()}.")
    else:
        st.success("No discrepancies — all cells score ≥ 90%!", icon="✅")

    # ── Download ─────────────────────────────────────────────────────────────
    st.markdown("---")
    with st.spinner("Building comparison workbook…"):
        acc_bytes = build_accuracy_excel(a_clean, a_ref, matched, field_data)

    st.download_button(
        "↓  Download Full Accuracy Report (.xlsx)  —  3 sheets: Summary · Row Detail · Discrepancies",
        data=acc_bytes,
        file_name="sov_accuracy_comparison.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True)


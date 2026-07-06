"""
Auto-extracted module: ui/phase3_transform.py
"""

from __future__ import annotations

import pandas as pd
import streamlit as st
import openpyxl
from fuzzywuzzy import fuzz  # type: ignore
import traceback
import tempfile
import io

from feedback.row_feedback.store import save_rule
from header_mapping.rms_crosswalk import AIR_TO_RMS_CONSTRUCTION, AIR_TO_RMS_OCCUPANCY
from row_processing.construction import _save_construction_rule
from row_processing.eda import run_eda
from row_processing.export import run_value_transformation
from row_processing.occupancy import _save_occupancy_rule, looks_like_occupancy_text
from ui.common import _human_basis, _match_type_display, safe_join
from ui.row_feedback_tab import render_row_edit_panel

def build_enhanced_qa_bytes(p3, locked_schema, mappings, raw_df: pd.DataFrame) -> bytes:
    """
    Build an enhanced QA report that includes:
    - Original Summary sheet
    - Original Flag_Log sheet
    - Mapping_Summary sheet (per-column mapping decisions)
    - Transformation_Summary sheet (per-column change stats + reasons)
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font

    HDR_FILL = PatternFill("solid", fgColor="4472C4")
    HDR_FONT = Font(bold=True, color="FFFFFF")
    ERR_FILL = PatternFill("solid", fgColor="FFC7CE")
    WRN_FILL = PatternFill("solid", fgColor="FFEB9C")
    INF_FILL = PatternFill("solid", fgColor="DEEBF7")

    def _hdr(ws, cols):
        ws.append(cols)
        for cell in ws[ws.max_row]:
            cell.fill = HDR_FILL; cell.font = HDR_FONT

    def _autowidth(ws):
        for col_cells in ws.columns:
            w = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[openpyxl.utils.get_column_letter(
                col_cells[0].column)].width = min(w + 2, 55)

    cleaned_df = p3["cleaned_df"]
    flag_log   = p3["flag_log"]
    flag_df    = flag_log.to_dataframe()

    # Start from the existing QA bytes so we keep the original Summary + Flag_Log sheets
    wb = openpyxl.load_workbook(io.BytesIO(p3["_qa_bytes"]))

    # ── Mapping Summary ───────────────────────────────────────────────────────
    if "Mapping_Summary" in wb.sheetnames:
        del wb["Mapping_Summary"]
    ws_map = wb.create_sheet("Mapping_Summary")
    _hdr(ws_map, ["Output Field", "Source Column(s)", "Match Type", "Confidence %",
                   "Flag", "Why Matched"])
    for m in mappings:
        label, _, _ = _match_type_display(m.match_type)
        ws_map.append([
            m.output_col,
            safe_join(m.source_cols) if m.source_cols else "—",
            label,
            m.confidence if m.confidence else 0,
            m.flag or "OK",
            _human_basis(m),
        ])
    _autowidth(ws_map)

    # ── Transformation Summary ────────────────────────────────────────────────
    if "Transformation_Summary" in wb.sheetnames:
        del wb["Transformation_Summary"]
    ws_tx = wb.create_sheet("Transformation_Summary")
    _hdr(ws_tx, ["Output Field", "Source Column(s)", "Rule Applied",
                  "Rows Changed", "% Changed", "Flag Types", "Error Count", "Warning Count"])
    decisions = [d for d in locked_schema.decisions
                 if d.final_source and d.decision != "unavailable"]
    rules_df  = p3.get("rules_df")
    n_rows = min(len(raw_df), len(cleaned_df))
    for d in decisions:
        oc = d.output_col
        if oc not in cleaned_df.columns:
            continue
        all_srcs = d.final_source
        ch = 0
        for i in range(n_rows):
            raw_parts = [str(raw_df.iloc[i].get(sc, "") or "").strip()
                         for sc in all_srcs if sc in raw_df.columns]
            rv = safe_join([r for r in raw_parts if r and r not in ("nan", "None", "")])
            cv = str(cleaned_df.iloc[i].get(oc, "")).strip()
            if rv != cv:
                ch += 1
        # Rule name from rules_df (the definitive source — always populated)
        if rules_df is not None and oc in rules_df.columns:
            unique_rules = rules_df[oc].dropna().unique().tolist()
            rule_name = unique_rules[0] if unique_rules else "—"
        else:
            rule_name = "—"
        if not flag_df.empty and "output_col" in flag_df.columns:
            col_f = flag_df[flag_df["output_col"] == oc]
            ftypes   = safe_join(col_f["flag_type"].dropna().unique().tolist())    if not col_f.empty else "—"
            n_err    = int((col_f["severity"] == "error").sum())   if not col_f.empty else 0
            n_warn   = int((col_f["severity"] == "warning").sum()) if not col_f.empty else 0
        else:
            ftypes = "—"; n_err = n_warn = 0
        ws_tx.append([oc, safe_join(all_srcs), rule_name, ch,
                       f"{100*ch//max(n_rows,1)}%", ftypes, n_err, n_warn])
    _autowidth(ws_tx)

    # ── Colour the Flag_Log sheet if it exists ────────────────────────────────
    if "Flag_Log" in wb.sheetnames:
        ws_fl = wb["Flag_Log"]
        for row in ws_fl.iter_rows(min_row=2):
            sev = str(row[3].value or "").lower() if len(row) > 3 else ""
            fill = ERR_FILL if sev == "error" else (WRN_FILL if sev == "warning" else INF_FILL)
            for cell in row:
                cell.fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_eda_panel(sov, raw_df, locked_schema, system, precomputed=None):
    """
    EDA panel — renders pre-computed EDA results (passed via precomputed dict).
    precomputed: output of sov.run_eda(), cached in session_state.
    """
    eda_results = precomputed or {}
    if not eda_results:
        st.info("No profiling data available.", icon="ℹ️")
        return

    # ── Summary bar ──────────────────────────────────────────────────────────
    total      = len(eda_results)
    crit_cols  = [f for f, e in eda_results.items() if any(i[0] == "critical" for i in e.get("issues", []))]
    warn_cols  = [f for f, e in eda_results.items() if any(i[0] == "warning"  for i in e.get("issues", []))]
    clean_cols = total - len(set(crit_cols) | set(warn_cols))
    low_fill   = [f for f, e in eda_results.items() if e.get("fill_rate", 100) < 80]

    ec1, ec2, ec3, ec4 = st.columns(4)
    ec1.metric("Fields Profiled",    total)
    ec2.metric("Critical Issues",    len(crit_cols),  help="< 50% fill rate or future years")
    ec3.metric("Warnings",           len(warn_cols),  help="Negative values, outliers, wide ranges")
    ec4.metric("Low Fill (< 80%)",   len(low_fill))

    if crit_cols:
        st.error(f"⚠️ Critical data issues in: {', '.join(f'`{f}`' for f in crit_cols)}", icon="🚨")
    if warn_cols and not crit_cols:
        st.warning(f"Review warnings in: {', '.join(f'`{f}`' for f in warn_cols[:6])}" +
                   (f" +{len(warn_cols)-6} more" if len(warn_cols) > 6 else ""))

    st.markdown("---")

    # ── Per-field expandable cards ────────────────────────────────────────────
    # Group by severity for display order: critical first, then warnings, then clean
    def _field_severity(f):
        issues = eda_results[f].get("issues", [])
        if any(i[0] == "critical" for i in issues): return 0
        if any(i[0] in ("error","warning") for i in issues): return 1
        if any(i[0] == "info" for i in issues): return 2
        return 3

    sorted_fields = sorted(eda_results.keys(), key=_field_severity)

    # Quick filter
    filt = st.radio("Show fields", ["All", "Issues only", "Clean only"],
                    horizontal=True, key="eda_filter")

    for field in sorted_fields:
        eda = eda_results[field]
        issues = eda.get("issues", [])
        has_issue = any(i[0] in ("critical","error","warning") for i in issues)
        if filt == "Issues only" and not has_issue: continue
        if filt == "Clean only"  and has_issue:     continue

        fill_rate   = eda.get("fill_rate", 0)
        itype       = eda.get("inferred_type", "—")
        null_ct     = eda.get("null_count", 0)
        unique_ct   = eda.get("unique_count", 0)
        src_cols    = eda.get("source_cols", [])
        samples     = eda.get("sample_values", [])

        # Severity badge for expander label
        if any(i[0] == "critical" for i in issues):
            sev_icon = "🔴"
        elif any(i[0] in ("error","warning") for i in issues):
            sev_icon = "🟡"
        elif issues:
            sev_icon = "🔵"
        else:
            sev_icon = "🟢"

        fill_colour = ("#10b981" if fill_rate >= 80 else
                       "#f59e0b" if fill_rate >= 50 else "#ef4444")

        expanded = has_issue and fill_rate < 50
        label = (f"{sev_icon} **{field}** · "
                 f"source: `{', '.join(src_cols)}` · "
                 f"fill: {fill_rate}% · type: {itype}")

        with st.expander(label, expanded=expanded):
            col_l, col_r = st.columns([3, 2])

            with col_l:
                # Fill rate progress bar
                st.markdown(
                    f'<div style="margin-bottom:.5rem">'
                    f'<span style="font-size:.75rem;color:#6c757d;font-weight:600">FILL RATE</span>'
                    f'<div style="background:#e5e7eb;border-radius:4px;height:8px;margin-top:.25rem">'
                    f'<div style="background:{fill_colour};width:{fill_rate}%;height:8px;border-radius:4px"></div>'
                    f'</div>'
                    f'<span style="font-family:\'IBM Plex Mono\',monospace;font-size:.82rem;'
                    f'color:{fill_colour};font-weight:700">{fill_rate}%</span>'
                    f'<span style="font-size:.72rem;color:#9ca3af;margin-left:.5rem">'
                    f'({null_ct} nulls · {unique_ct} unique)</span>'
                    f'</div>', unsafe_allow_html=True)

                # Stats for numeric / year fields
                stat_pairs = []
                for k in ("min","max","mean","median"):
                    if k in eda:
                        v = eda[k]
                        stat_pairs.append(f"<b>{k}</b>: {v:,.1f}" if isinstance(v, float) else f"<b>{k}</b>: {v}")
                if stat_pairs:
                    st.markdown(
                        f'<div style="font-family:\'IBM Plex Mono\',monospace;font-size:.78rem;'
                        f'background:#f0f4f8;padding:.4rem .6rem;border-radius:4px;margin:.3rem 0">'
                        + "  ·  ".join(stat_pairs) +
                        f'</div>', unsafe_allow_html=True)

                # Top values for text fields
                top_vals = eda.get("top_values", {})
                if top_vals:
                    tv_rows = "".join(
                        f'<div style="display:flex;justify-content:space-between;font-size:.73rem;'
                        f'padding:.12rem 0;border-bottom:1px solid #f3f4f6">'
                        f'<span style="font-family:\'IBM Plex Mono\',monospace;color:#1a1a2e">{v}</span>'
                        f'<span style="color:#6c757d">{cnt}×</span></div>'
                        for v, cnt in list(top_vals.items())[:5])
                    st.markdown(
                        f'<div style="margin:.3rem 0">'
                        f'<div style="font-size:.68rem;color:#6c757d;font-weight:600;'
                        f'text-transform:uppercase;letter-spacing:.06em;margin-bottom:.2rem">'
                        f'Top values</div>{tv_rows}</div>', unsafe_allow_html=True)

                # Outliers
                if eda.get("outlier_count"):
                    outlier_vals = eda.get("outlier_values", [])
                    st.markdown(
                        f'<div style="background:#fffbf0;border:1px solid #f59e0b40;'
                        f'border-left:3px solid #f59e0b;border-radius:0 4px 4px 0;'
                        f'padding:.3rem .6rem;font-size:.75rem;margin:.3rem 0">'
                        f'<b>⚡ {eda["outlier_count"]} outliers</b>'
                        + (f': {", ".join(str(int(v)) for v in outlier_vals[:5])}' if outlier_vals else "")
                        + f'</div>', unsafe_allow_html=True)

            with col_r:
                # Issues list
                if issues:
                    issue_html = ""
                    sev_styles = {
                        "critical": ("#fef2f2", "#dc2626", "🔴"),
                        "error":    ("#fef2f2", "#dc2626", "🔴"),
                        "warning":  ("#fffbf0", "#b45309", "🟡"),
                        "info":     ("#eff6ff", "#1d4ed8", "🔵"),
                    }
                    for sev, msg in issues:
                        bg, fg, icon = sev_styles.get(sev, ("#f9fafb", "#374151", "⚪"))
                        issue_html += (
                            f'<div style="background:{bg};border-radius:4px;padding:.25rem .5rem;'
                            f'margin-bottom:.2rem;font-size:.74rem;color:{fg}">'
                            f'{icon} {msg}</div>')
                    st.markdown(
                        f'<div style="margin-bottom:.5rem">'
                        f'<div style="font-size:.68rem;color:#6c757d;font-weight:600;'
                        f'text-transform:uppercase;letter-spacing:.06em;margin-bottom:.3rem">'
                        f'Data Quality Flags</div>{issue_html}</div>', unsafe_allow_html=True)

                # Sample values
                if samples:
                    samp_html = "  ·  ".join(
                        f'<code style="font-size:.72rem;background:#e8f0fe;color:#1565c0;'
                        f'padding:.05rem .3rem;border-radius:2px">{s[:30]}</code>'
                        for s in samples[:6])
                    st.markdown(
                        f'<div style="margin-top:.3rem">'
                        f'<div style="font-size:.68rem;color:#6c757d;font-weight:600;'
                        f'text-transform:uppercase;letter-spacing:.06em;margin-bottom:.3rem">Samples</div>'
                        f'{samp_html}</div>', unsafe_allow_html=True)

                # Address-specific multi-address warning
                if field == "Street" and eda.get("inferred_type") == "text":
                    multi_addr_ct = sum(
                        1 for v in samples
                        if any(sep in v for sep in ["\n", ";", " / ", " & ", "|"])
                        or len([p for p in v.split(",") if p.strip()]) > 2
                    )
                    if multi_addr_ct:
                        st.markdown(
                            f'<div style="background:#fef3c7;border-left:3px solid #f59e0b;'
                            f'padding:.3rem .5rem;font-size:.74rem;color:#92400e;margin-top:.3rem;'
                            f'border-radius:0 4px 4px 0">'
                            f'⚠️ {multi_addr_ct} of {len(samples)} samples appear to contain multiple addresses. '
                            f'Pipeline keeps first address only.</div>', unsafe_allow_html=True)


def build_cleaned_with_reasons(
    cleaned_df: pd.DataFrame,
    raw_df: pd.DataFrame,
    locked_schema,
    flag_log,
    rules_df=None,
) -> bytes:
    """
    Build a cleaned SOV workbook that includes a 'Reason' row below each
    data row explaining what transformation was applied to each field.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    HDR_FILL  = PatternFill("solid", fgColor="1B3A6B")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
    RSN_FILL  = PatternFill("solid", fgColor="EFF6FF")
    RSN_FONT  = Font(italic=True, color="6B7280", size=8, name="Calibri")
    CHG_FILL  = PatternFill("solid", fgColor="FEF3C7")
    ERR_FILL  = PatternFill("solid", fgColor="FEE2E2")

    # Build flag lookup: (row_idx, output_col) -> list of messages
    flag_df = flag_log.to_dataframe() if flag_log is not None else pd.DataFrame()
    flag_lookup: dict = {}
    if not flag_df.empty and "row_idx" in flag_df.columns and "output_col" in flag_df.columns:
        for _, frow in flag_df.iterrows():
            key = (int(frow["row_idx"]), str(frow["output_col"]))
            flag_lookup.setdefault(key, []).append(
                str(frow.get("message", frow.get("flag_type", "flagged")))
            )

    # Build decision lookup: output_col -> final_source list
    dec_lookup: dict = {}
    if locked_schema is not None:
        for d in locked_schema.decisions:
            dec_lookup[d.output_col] = d.final_source or []

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Cleaned_With_Reasons"

    # ── Build interleaved column order: RAW src | TARGET | RAW src | TARGET … ──
    RAW_FILL  = openpyxl.styles.PatternFill("solid", fgColor="D9E1F2")
    RAW_FONT  = openpyxl.styles.Font(bold=True, color="333333", size=9, name="Calibri", italic=True)
    SRC_FILL  = openpyxl.styles.PatternFill("solid", fgColor="EBF3FB")  # light blue for raw data cells
    RSN2_HDR  = openpyxl.styles.PatternFill("solid", fgColor="F0F0F0")

    # interleaved: list of (header_label, col_name, is_raw, src_col_name_or_None)
    interleaved = []
    cols = list(cleaned_df.columns)
    for col in cols:
        src_cols_for_col = dec_lookup.get(col, [])
        for sc in src_cols_for_col:
            if sc in raw_df.columns:
                interleaved.append((f"RAW: {sc}", sc, True, sc))
        interleaved.append((col, col, False, None))

    n = len(cleaned_df)
    total_cols = len(interleaved)

    # ── Header row ────────────────────────────────────────────────────────────
    for ci, (label, _, is_raw, _sc) in enumerate(interleaved, 1):
        cell = ws.cell(1, ci, label)
        cell.fill = RAW_FILL if is_raw else HDR_FILL
        cell.font = RAW_FONT if is_raw else HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Data + reason rows ────────────────────────────────────────────────────
    for ri in range(n):
        data_row   = ri * 2 + 2
        reason_row = ri * 2 + 3

        for ci, (label, col_name, is_raw, src_col) in enumerate(interleaved, 1):
            if is_raw:
                # Write raw source value
                rv = str(raw_df.iloc[ri].get(src_col, "") if src_col in raw_df.columns else "").strip()
                rv = "" if rv in ("nan", "None") else rv
                cell = ws.cell(data_row, ci, rv)
                cell.fill = SRC_FILL
                cell.font = Font(size=9, name="Calibri", italic=True)
                cell.alignment = Alignment(vertical="center")
                # Reason cell for raw column — show col name
                rc = ws.cell(reason_row, ci, f"Source: {src_col}")
                rc.fill = RSN_FILL; rc.font = RSN_FONT
                rc.alignment = Alignment(vertical="center")
            else:
                col = col_name
                cleaned_val = cleaned_df.iloc[ri].get(col, "")

                src_cols_for_col = dec_lookup.get(col, [])
                raw_parts = []
                for sc in src_cols_for_col:
                    rv2 = str(raw_df.iloc[ri].get(sc, "") if sc in raw_df.columns else "").strip()
                    if rv2 and rv2 not in ("nan", "None", ""):
                        raw_parts.append(rv2)
                raw_val = safe_join(raw_parts)
                changed = raw_val != str(cleaned_val).strip()

                flags = flag_lookup.get((ri, col), [])

                # openpyxl cannot handle pd.NA/pd.NaT/numpy NA
                try:
                    import pandas as _pd_chk
                    if _pd_chk.isna(cleaned_val):
                        cleaned_val = ""
                except (TypeError, ValueError):
                    pass
                if cleaned_val is None:
                    cleaned_val = ""
                dc_cell = ws.cell(data_row, ci, cleaned_val)
                dc_cell.font = Font(size=9, name="Calibri", bold=True)
                dc_cell.alignment = Alignment(vertical="center")
                if flags:
                    dc_cell.fill = ERR_FILL
                elif changed:
                    dc_cell.fill = CHG_FILL

                if flags:
                    reason_str = "; ".join(flags[:2])
                elif rules_df is not None and col in rules_df.columns:
                    unique_rules = rules_df[col].dropna().unique().tolist()
                    reason_str = unique_rules[0] if unique_rules else ("Changed" if changed else "No change")
                elif changed:
                    reason_str = f"Transformed from: {raw_val[:60]}" if raw_val else "Derived"
                else:
                    reason_str = "No change"

                rc = ws.cell(reason_row, ci, reason_str)
                rc.fill = RSN_FILL; rc.font = RSN_FONT
                rc.alignment = Alignment(vertical="center", wrap_text=True)

        ws.row_dimensions[reason_row].height = 14

    # ── Auto-width ────────────────────────────────────────────────────────────
    for col_cells in ws.columns:
        w = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_cells[0].column)
        ].width = min(w + 2, 30)

    ws.freeze_panes = "B2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def render_code_review_panel(p3, sov, *, field_prefix, code_field, code_map,
                              save_rule_fn=None, code_filter=None, panel_key=""):
    """
    Generic grouped review panel for any AI/semantic-mapped code field.
    Groups rows by their SUGGESTED code so you click a code and see every
    row mapped to it, instead of scrolling a flat random-order list.

    field_prefix : "_occ" or "_constr"  (matches the helper columns)
    code_field   : "OccupancyCode" or "ConstructionCode"
    code_map     : sov.AIR_TO_RMS_OCCUPANCY / sov.AIR_TO_RMS_CONSTRUCTION
    save_rule_fn : sov._save_occupancy_rule / sov._save_construction_rule
    code_filter  : optional fn(code)->bool to restrict the override dropdown
    """
    cleaned_df = p3["cleaned_df"]
    review_col = f"{field_prefix}_needs_review"
    method_col = f"{field_prefix}_method"
    conf_col   = f"{field_prefix}_confidence"
    raw_col    = f"{field_prefix}_raw_description"

    if review_col not in cleaned_df.columns:
        return

    review_df = cleaned_df[cleaned_df[review_col] == True].copy()
    if review_df.empty:
        st.success(
            f"All {code_field} values mapped with high confidence — no review needed.",
            icon="✅")
        return

    st.markdown(
        f'<div style="background:rgba(239,68,68,.07);border:1px solid rgba(239,68,68,.3);'
        f'border-radius:6px;padding:.7rem 1rem;margin-bottom:1rem">'
        f'<b style="color:#dc2626">⚠️ {len(review_df)} row(s) need {code_field} review</b> — '
        f'mapped by semantic or AI matching with low confidence. '
        f'Grouped by suggested code — expand a code to see its rows.</div>',
        unsafe_allow_html=True)

    code_options = {
        f"{code} — {info['rms_label']}": code
        for code, info in sorted(code_map.items())
        if code_filter is None or code_filter(code)
    }
    code_labels = list(code_options.keys())

    def _group_key(v):
        try:
            return int(v)
        except (TypeError, ValueError):
            return None

    review_df["_group_code"] = review_df[code_field].map(_group_key)
    groups = sorted(review_df.groupby("_group_code", dropna=False),
                     key=lambda kv: -len(kv[1]))   # biggest groups first

    confirmed_overrides = {}   # row_idx -> confirmed code

    for group_code, group_df in groups:
        rms_label = code_map.get(group_code, {}).get("rms_label", "Unrecognized") if group_code is not None else "Unrecognized / No Match"
        avg_conf  = round(group_df[conf_col].astype(float).mean()) if conf_col in group_df else 0
        header = (f"🏷️ AIR {group_code if group_code is not None else '—'} — {rms_label}"
                  f"  ·  {len(group_df)} row(s)  ·  avg confidence {avg_conf}%")

        with st.expander(header, expanded=False):
            # Real bulk override — applies immediately, no need to touch every row
            b1, b2 = st.columns([3, 2])
            with b1:
                bulk_choice = st.selectbox(
                    "Bulk-correct this entire group",
                    ["— keep individual choices —"] + code_labels,
                    key=f"{panel_key}_bulk_sel_{group_code}")
            with b2:
                st.write("")
                if bulk_choice != "— keep individual choices —" and st.button(
                        f"Apply to all {len(group_df)} rows",
                        key=f"{panel_key}_bulk_btn_{group_code}"):
                    chosen_code = code_options[bulk_choice]
                    for row_idx in group_df.index:
                        p3["cleaned_df"].at[row_idx, code_field] = chosen_code
                        p3["cleaned_df"].at[row_idx, review_col] = False
                        if save_rule_fn:
                            save_rule_fn(raw_description=str(group_df.loc[row_idx, raw_col]),
                                         confirmed_code=chosen_code)
                    st.success(f"✅ {len(group_df)} row(s) set to {bulk_choice}", icon="✅")
                    st.rerun()

            st.markdown("---")

            for row_idx, row in group_df.iterrows():
                raw_desc  = str(row.get(raw_col, ""))
                suggested = row.get(code_field)
                method    = row.get(method_col, "unknown")
                conf      = row.get(conf_col, 0)
                loc_name  = str(row.get("LocationName", f"Row {row_idx + 1}"))

                sugg_label = next((lbl for lbl, c in code_options.items() if c == suggested),
                                   code_labels[0] if code_labels else None)
                conf_color = "#10b981" if conf >= 75 else "#f59e0b" if conf >= 50 else "#ef4444"
                method_label = {"semantic": "Semantic match", "ai": "AI inference",
                                 "confirmed_rule": "Confirmed rule"}.get(method, method)

                with st.expander(
                    f"📍 {loc_name} — '{raw_desc[:60]}{'…' if len(raw_desc) > 60 else ''}'",
                    expanded=False
                ):
                    col_l, col_r = st.columns([3, 2])
                    with col_l:
                        st.markdown(
                            f'**Original description:** `{raw_desc}`  \n'
                            f'**Suggested code:** `{suggested}` '
                            f'({code_map.get(int(suggested), {}).get("rms_label","?") if suggested else "None"})  \n'
                            f'**Matched by:** {method_label}  \n'
                            f'**Confidence:** '
                            f'<span style="color:{conf_color};font-weight:700">{conf}%</span>',
                            unsafe_allow_html=True)

                    with col_r:
                        action = st.radio(
                            "Action", ["Accept suggestion", "Override"],
                            key=f"{panel_key}_action_{row_idx}", horizontal=True)

                        if action == "Override" and code_labels:
                            chosen_label = st.selectbox(
                                f"Correct {code_field}", code_labels,
                                index=code_labels.index(sugg_label) if sugg_label in code_labels else 0,
                                key=f"{panel_key}_override_{row_idx}")
                            final_code = code_options[chosen_label]
                        else:
                            final_code = int(suggested) if suggested else None

                        # Soft sanity check: does this description actually
                        # carry use-type/construction-type signal, or is it
                        # just a bare proper-noun building name? If the latter,
                        # default the "remember" checkbox OFF so a one-off
                        # correction for THIS location doesn't silently get
                        # cached as a global rule applied to every future row
                        # that happens to share the same text.
                        _has_signal = True
                        try:
                            if field_prefix == "_occ":
                                _has_signal = sov.looks_like_occupancy_text(raw_desc)
                            elif field_prefix == "_constr":
                                from feedback.construction_aliases import _looks_like_construction_text
                                _has_signal = _looks_like_construction_text(raw_desc)
                        except Exception:
                            _has_signal = True  # fail open, don't block the UI

                        save_rule = st.checkbox(
                            "Remember this mapping for future uploads",
                            value=_has_signal,
                            key=f"{panel_key}_save_{row_idx}", disabled=save_rule_fn is None)
                        if not _has_signal:
                            st.caption(
                                "⚠️ This description doesn't clearly name a use/construction "
                                "type — it may be a location-specific name. Left unchecked so "
                                "it won't be applied to unrelated rows in future uploads.")

                        confirmed_overrides[row_idx] = {
                            "code": final_code, "raw_desc": raw_desc,
                            "save_rule": save_rule, "action": action,
                        }

    if confirmed_overrides and st.button(
        "✅  Apply Confirmed Codes", use_container_width=True, key=f"{panel_key}_apply_btn"
    ):
        for row_idx, decision in confirmed_overrides.items():
            p3["cleaned_df"].at[row_idx, code_field] = decision["code"]
            p3["cleaned_df"].at[row_idx, review_col] = False

        rules_saved = 0
        if save_rule_fn:
            for row_idx, decision in confirmed_overrides.items():
                if not decision["save_rule"] or not decision["raw_desc"]:
                    continue
                save_rule_fn(raw_description=decision["raw_desc"], confirmed_code=decision["code"])
                rules_saved += 1

        st.success(
            f"✅ {len(confirmed_overrides)} code(s) confirmed"
            + (f" · {rules_saved} rule(s) saved for future uploads" if rules_saved else ""),
            icon="✅")
        st.rerun()


def render_code_browser_panel(p3, sov, *, field_prefix, code_field, code_map,
                               save_rule_fn=None, code_filter=None, panel_key=""):
    """
    Free-form search/browse panel: lets the analyst look up a description
    (or a code) and see EVERY row currently mapped to it — not just the
    flagged/low-confidence subset that render_code_review_panel shows.

    field_prefix : "_occ" or "_constr"  (matches the helper columns)
    code_field   : "OccupancyCode" or "ConstructionCode"
    code_map     : sov.AIR_TO_RMS_OCCUPANCY / sov.AIR_TO_RMS_CONSTRUCTION
    save_rule_fn : sov._save_occupancy_rule / sov._save_construction_rule
    code_filter  : optional fn(code)->bool to restrict the override dropdown
    """
    cleaned_df = p3["cleaned_df"]
    review_col = f"{field_prefix}_needs_review"
    method_col = f"{field_prefix}_method"
    conf_col   = f"{field_prefix}_confidence"
    raw_col    = f"{field_prefix}_raw_description"

    if raw_col not in cleaned_df.columns or code_field not in cleaned_df.columns:
        return

    st.markdown("##### 🔍 Search / browse all mapped rows")
    st.caption(
        "Look up every row assigned to a description or code — including "
        "ones that were auto-mapped with high confidence and never flagged.")

    code_options = {
        f"{code} — {info['rms_label']}": code
        for code, info in sorted(code_map.items())
        if code_filter is None or code_filter(code)
    }
    code_labels = list(code_options.keys())

    search_mode = st.radio(
        "Search by", ["Description", f"Assigned {code_field}"],
        key=f"{panel_key}_browse_mode", horizontal=True)

    match_df = None

    if search_mode == "Description":
        desc_counts = cleaned_df[raw_col].dropna().astype(str)
        desc_counts = desc_counts[desc_counts.str.strip() != ""].value_counts()
        all_descs   = list(desc_counts.items())   # [(desc, row_count), ...]

        MAX_DROPDOWN = 60   # keep the dropdown light with hundreds of SOVs loaded

        search_term = st.text_input(
            "Search a description (partial words ok — e.g. 'rice' matches "
            "'Rice Mill', 'Rice Mill 2')",
            key=f"{panel_key}_browse_desc_search", placeholder="Start typing…")

        if search_term.strip():
            term = search_term.strip().lower()
            scored = []
            for d, n in all_descs:
                d_low = d.lower()
                if term in d_low:
                    score = 100   # exact substring hit — always ranks first
                else:
                    score = max(
                        sov.fuzz.partial_ratio(term, d_low),
                        sov.fuzz.token_set_ratio(term, d_low),
                    ) if hasattr(sov, "fuzz") else 0
                if score >= 60:
                    scored.append((d, n, score))
            scored.sort(key=lambda t: (-t[2], -t[1]))
            filtered = [(d, n) for d, n, _ in scored]
            truncated = len(filtered) > MAX_DROPDOWN
            filtered = filtered[:MAX_DROPDOWN]
        else:
            filtered = sorted(all_descs, key=lambda t: -t[1])
            truncated = len(filtered) > MAX_DROPDOWN
            filtered = filtered[:MAX_DROPDOWN]

        if not filtered:
            st.info("No descriptions match that search.", icon="ℹ️")
            return

        if truncated:
            st.caption(
                f"Showing top {MAX_DROPDOWN} matches — refine your search to narrow further.")

        desc_display_to_raw = {
            f"{d}  ({n} row{'s' if n != 1 else ''})": d for d, n in filtered
        }

        combine_all = False
        if search_term.strip() and len(filtered) > 1:
            combine_all = st.checkbox(
                f"Combine all {len(filtered)} matching descriptions into one view "
                "(e.g. treat 'Rice Mill' and 'Rice Mill 2' as one group)",
                key=f"{panel_key}_browse_combine")

        if combine_all:
            chosen_descs = list(desc_display_to_raw.values())
            match_df = cleaned_df[cleaned_df[raw_col].astype(str).isin(chosen_descs)]
        else:
            picked = st.selectbox(
                "Pick the exact description",
                ["— select —"] + list(desc_display_to_raw.keys()),
                key=f"{panel_key}_browse_desc")
            if picked != "— select —":
                chosen_desc = desc_display_to_raw[picked]
                match_df = cleaned_df[cleaned_df[raw_col].astype(str) == chosen_desc]
    else:
        UNMAPPED_LABEL = "⚠️ — Unmapped / Not Identified (blank code) —"
        picked = st.selectbox(
            f"Select a {code_field}", ["— select —", UNMAPPED_LABEL] + code_labels,
            key=f"{panel_key}_browse_code")
        if picked == UNMAPPED_LABEL:
            match_df = cleaned_df[cleaned_df[code_field].isna()]
        elif picked != "— select —":
            chosen_code = code_options[picked]
            match_df = cleaned_df[cleaned_df[code_field] == chosen_code]

    if match_df is None:
        return
    if match_df.empty:
        st.info("No rows found for that search.", icon="ℹ️")
        return

    st.markdown(f"**{len(match_df)} row(s) found**")

    display_cols = [c for c in
                     ["LocationName", raw_col, code_field, method_col, conf_col, review_col]
                     if c in match_df.columns]
    rename_map = {raw_col: "Raw Description", code_field: "Assigned Code",
                  method_col: "Method", conf_col: "Confidence %",
                  review_col: "Flagged for review"}
    st.dataframe(match_df[display_cols].rename(columns=rename_map),
                 use_container_width=True, hide_index=True)

    # Bulk re-assign every row currently shown by the search/filter above
    b1, b2 = st.columns([3, 2])
    with b1:
        bulk_choice = st.selectbox(
            "Bulk-correct all matched rows above",
            ["— keep individual choices —"] + code_labels,
            key=f"{panel_key}_browse_bulk_sel")
    with b2:
        st.write("")
        if bulk_choice != "— keep individual choices —" and st.button(
                f"Apply to all {len(match_df)} rows",
                key=f"{panel_key}_browse_bulk_btn"):
            chosen_code = code_options[bulk_choice]
            for row_idx in match_df.index:
                p3["cleaned_df"].at[row_idx, code_field] = chosen_code
                if review_col in p3["cleaned_df"].columns:
                    p3["cleaned_df"].at[row_idx, review_col] = False
                if save_rule_fn:
                    save_rule_fn(raw_description=str(match_df.loc[row_idx, raw_col]),
                                 confirmed_code=chosen_code)
            st.success(f"✅ {len(match_df)} row(s) set to {bulk_choice}", icon="✅")
            st.rerun()

    with st.expander(f"Edit individual rows ({len(match_df)})", expanded=False):
        for row_idx, row in match_df.iterrows():
            loc_name  = str(row.get("LocationName", f"Row {row_idx + 1}"))
            current   = row.get(code_field)
            cur_label = next((lbl for lbl, c in code_options.items() if c == current), None)

            r1, r2 = st.columns([3, 2])
            with r1:
                st.markdown(f"**{loc_name}** — current: `{current}`  \n"
                             f"*{row.get(raw_col, '')}*")
            with r2:
                new_label = st.selectbox(
                    "Correct code", code_labels,
                    index=code_labels.index(cur_label) if cur_label in code_labels else 0,
                    key=f"{panel_key}_browse_row_{row_idx}", label_visibility="collapsed")
                if st.button("Save", key=f"{panel_key}_browse_row_btn_{row_idx}"):
                    new_code = code_options[new_label]
                    p3["cleaned_df"].at[row_idx, code_field] = new_code
                    if review_col in p3["cleaned_df"].columns:
                        p3["cleaned_df"].at[row_idx, review_col] = False
                    if save_rule_fn:
                        save_rule_fn(raw_description=str(row.get(raw_col, "")),
                                     confirmed_code=new_code)
                    st.success(f"Updated {loc_name} → {new_label}", icon="✅")
                    st.rerun()
            st.markdown("---")


def render_occupancy_review_panel(p3, sov):
    render_code_review_panel(
        p3, sov, field_prefix="_occ", code_field="OccupancyCode",
        code_map=sov.AIR_TO_RMS_OCCUPANCY, save_rule_fn=sov._save_occupancy_rule,
        code_filter=lambda c: c < 400, panel_key="occ")
    st.markdown("---")
    render_code_browser_panel(
        p3, sov, field_prefix="_occ", code_field="OccupancyCode",
        code_map=sov.AIR_TO_RMS_OCCUPANCY, save_rule_fn=sov._save_occupancy_rule,
        code_filter=lambda c: c < 400, panel_key="occ")


def render_construction_review_panel(p3, sov):
    render_code_review_panel(
        p3, sov, field_prefix="_constr", code_field="ConstructionCode",
        code_map=sov.AIR_TO_RMS_CONSTRUCTION,
        save_rule_fn=getattr(sov, "_save_construction_rule", None),
        panel_key="constr")
    st.markdown("---")
    render_code_browser_panel(
        p3, sov, field_prefix="_constr", code_field="ConstructionCode",
        code_map=sov.AIR_TO_RMS_CONSTRUCTION,
        save_rule_fn=getattr(sov, "_save_construction_rule", None),
        panel_key="constr")


def render_phase3(sov, system):
    if "locked_schema" not in st.session_state or "phase1_result" not in st.session_state:
        st.info("Complete Phases 1 and 2 first.", icon="ℹ️")
        return

    st.markdown("Apply field-level transformation rules automatically: address formatting, country ISO resolution, construction/occupancy code lookup, and postal formatting.")

    raw_df     = st.session_state["phase1_result"]["data_frame"]
    locked_sch = st.session_state["locked_schema"]

    # ── EDA Panel — auto-runs once on schema lock, results cached ───────────
    eda_cache_key = "eda_results"   # single key; re-lock clears it via the button below
    eda_error_key = "eda_error"

    # Clear stale EDA if schema was re-locked (review_timestamp changes)
    eda_ts_key = "eda_locked_ts"
    current_ts = getattr(locked_sch, "review_timestamp", "")
    if st.session_state.get(eda_ts_key) != current_ts:
        st.session_state.pop(eda_cache_key, None)
        st.session_state.pop(eda_error_key, None)
        st.session_state[eda_ts_key] = current_ts

    if eda_cache_key not in st.session_state:
        with st.spinner("Analysing source data…"):
            try:
                eda_results = sov.run_eda(raw_df, locked_sch, target_system=system)
                st.session_state[eda_cache_key] = eda_results
                st.session_state.pop(eda_error_key, None)
            except Exception as _eda_exc:
                import traceback as _tb
                st.session_state[eda_cache_key] = None
                st.session_state[eda_error_key] = f"{_eda_exc}\n\n{_tb.format_exc()}"

    with st.expander("🔍  Source Data Analysis", expanded=True):
        st.markdown(
            '<p style="font-size:.82rem;color:#6c757d;margin-bottom:.8rem">'
            'Profiled each mapped source column automatically on schema lock — '
            'outliers, low fill rates, mixed types, multi-address rows, and future-year values.</p>',
            unsafe_allow_html=True)

        if st.session_state.get(eda_error_key):
            st.error("EDA failed — see details below.", icon="🚨")
            with st.expander("EDA error traceback"):
                st.code(st.session_state[eda_error_key])
        else:
            cached_eda = st.session_state.get(eda_cache_key) or {}
            if cached_eda:
                _render_eda_panel(sov, raw_df, locked_sch, system, precomputed=cached_eda)
            else:
                st.info("No mapped source columns to profile.", icon="ℹ️")

    st.markdown("---")

    if st.button("⚡  Run Transformation", use_container_width=True):
        with st.spinner("Applying transformation rules…"):
            try:
                with tempfile.TemporaryDirectory() as tmpdir:
                    p3 = sov.run_value_transformation(
                        phase1_result=st.session_state["phase1_result"],
                        locked_schema=st.session_state["locked_schema"],
                        target_system=system, output_dir=tmpdir,
                        report_name="sov_cleaned", lob_col="",
                        apply_proration=False, proration_group_col=None)
                    with open(p3["output_excel"],"rb") as f: p3["_excel_bytes"] = f.read()
                    with open(p3["output_json"], "r") as f: p3["_json_str"]   = f.read()
                    with open(p3["qa_report"],   "rb") as f: p3["_qa_bytes"]  = f.read()
                st.session_state["phase3_result"] = p3
                st.success("Transformation complete!", icon="⚡")
            except Exception as e:
                st.error(f"Transformation failed: {e}")
                with st.expander("Traceback"): st.code(traceback.format_exc())
                return

    if "phase3_result" not in st.session_state:
        return

    p3 = st.session_state["phase3_result"]
    cleaned_df = p3["cleaned_df"]; flag_log = p3["flag_log"]
    flag_df = flag_log.to_dataframe(); total_flags = sum(p3["flag_summary"].values())
    flagged_rows = flag_df["row_idx"].nunique() if not flag_df.empty else 0

    st.markdown("---")
    rule_app_log = p3.get("rule_application_log", [])
    n_rule_cells = len(rule_app_log)
    c1,c2,c3,c4,c5 = st.columns(5)
    c1.metric("Rows Processed",    len(cleaned_df))
    c2.metric("Total Flags",      total_flags)
    c3.metric("Rows with Flags",  flagged_rows)
    c4.metric("Clean Rows",       len(cleaned_df) - flagged_rows)
    c5.metric("Rule Overrides",   n_rule_cells, help="Cells changed by row feedback rules")
    if n_rule_cells:
        st.markdown(
            f'<div style="background:rgba(16,185,129,.08);border:1px solid rgba(16,185,129,.3);'
            f'border-radius:6px;padding:.5rem 1rem;display:inline-flex;align-items:center;'
            f'gap:.5rem;margin-top:.3rem">'
            f'<span style="font-size:.9rem">⚙️</span>'
            f'<span style="font-family:var(--mono);font-size:.82rem;color:#065f46;font-weight:600">'
            f'{n_rule_cells} cell(s) corrected by saved row rules</span>'
            f'</div>',
            unsafe_allow_html=True)

    tabs3 = st.tabs(["Before vs After", "Cleaned Output", "Flag Log"])

    with tabs3[0]:
        raw_df = st.session_state["phase1_result"]["data_frame"]
        locked = st.session_state["locked_schema"]
        mapped_fields = [d for d in locked.decisions if d.final_source and d.decision != "unavailable"]
        field_names   = [d.output_col for d in mapped_fields if d.output_col in cleaned_df.columns]
        if field_names:
            chosen = st.selectbox("Select a field to inspect", field_names, key="bva_field")
            dec = next((d for d in mapped_fields if d.output_col == chosen), None)
            if dec and chosen in cleaned_df.columns:
                all_srcs = dec.final_source
                n = min(len(raw_df), len(cleaned_df))
                bva = []
                for i in range(n):
                    raw_parts = []
                    for src_c in all_srcs:
                        rv = str(raw_df.iloc[i].get(src_c, "") if src_c in raw_df.columns else "").strip()
                        if rv and rv not in ("nan","None",""):
                            raw_parts.append(f"{src_c}: {rv}" if len(all_srcs) > 1 else rv)
                    rv_display = safe_join(raw_parts) if raw_parts else ""
                    cv = str(cleaned_df.iloc[i].get(chosen,"")).strip()
                    bva.append({"Row": i+1, "Original Value": rv_display, "Cleaned Value": cv,
                                "Changed": "✏" if rv_display != cv else "="})
                bva_df = pd.DataFrame(bva)
                def sty_ch(v): return "color:#f59e0b;font-weight:bold" if v=="✏" else "color:#10b981"
                st.dataframe(bva_df.style.map(sty_ch,subset=["Changed"]),
                             use_container_width=True, height=min(40*n+60,500))
                ch = sum(1 for r in bva if r["Changed"]=="✏")
                s1,s2,s3 = st.columns(3)
                s1.metric("Total Rows",n); s2.metric("Values Changed",ch); s3.metric("Unchanged",n-ch)
                if len(all_srcs) > 1:
                    st.info(f"This field combines {len(all_srcs)} source columns: {safe_join(f'`{s}`' for s in all_srcs)}", icon="ℹ️")
        else:
            st.info("No mapped fields with source columns to compare.")

        st.markdown("---"); st.markdown("**Change summary across all fields**")
        sum_rows = []
        for d in mapped_fields:
            oc = d.output_col
            if oc not in cleaned_df.columns: continue
            n = min(len(raw_df), len(cleaned_df))
            all_srcs = d.final_source
            ch = 0
            for i in range(n):
                raw_parts = []
                for src_c in all_srcs:
                    rv = str(raw_df.iloc[i].get(src_c, "") if src_c in raw_df.columns else "").strip()
                    if rv and rv not in ("nan","None",""): raw_parts.append(rv)
                rv_combined = safe_join(raw_parts)
                cv = str(cleaned_df.iloc[i].get(oc,"")).strip()
                if rv_combined != cv: ch += 1
            src_label = safe_join(all_srcs) if all_srcs else "—"
            rules_df = p3.get("rules_df")
            if rules_df is not None and oc in rules_df.columns:
                unique_rules = rules_df[oc].dropna().unique().tolist()
                reason_str = unique_rules[0] if unique_rules else "—"
            else:
                reason_str = "—"
            sum_rows.append({"Output Field": oc, "Source Column(s)": src_label,
                              "Rows Changed": ch, "Total Rows": n,
                              "% Changed": f"{100*ch//max(n,1)}%",
                              "Reason": reason_str})
        if sum_rows:
            sdf = pd.DataFrame(sum_rows).sort_values("Rows Changed", ascending=False)
            def sty_p(v):
                try:
                    p = int(v.replace("%",""))
                    if p>50: return "color:#ef4444;font-weight:bold"
                    if p>0:  return "color:#f59e0b"
                    return "color:#10b981"
                except Exception: return ""
            st.dataframe(sdf.style.map(sty_p,subset=["% Changed"]),
                         use_container_width=True, height=380)

        
        # Occupancy review panel
        st.markdown("---")
        st.markdown("### 🏢  Occupancy Code Review")
        render_occupancy_review_panel(p3, sov)

        # Construction review panel
        st.markdown("---")
        st.markdown("### 🧱  Construction Code Review")
        render_construction_review_panel(p3, sov)        


        # ── Cell editor for row-level feedback rules ──────────────────────
        render_row_edit_panel(p3, system)

    with tabs3[1]:
        _display_df = cleaned_df[[c for c in cleaned_df.columns
                               if not c.startswith("_occ_") and not c.startswith("_constr_")]]
        st.markdown(f'<p style="font-size:.82rem;color:var(--muted)">{len(cleaned_df)} rows × {len(cleaned_df.columns)} columns</p>', unsafe_allow_html=True)
        st.dataframe(_display_df.head(50), use_container_width=True, height=420)

    with tabs3[2]:
        if flag_df.empty:
            st.success("No flags — all rows transformed cleanly!", icon="✅")
        else:
            st.markdown(f'<p style="font-size:.82rem;color:var(--muted)">{len(flag_df)} flag entries across {flagged_rows} rows</p>', unsafe_allow_html=True)
            def sty_s(v):
                if v=="error": return "color:#ef4444;font-weight:bold"
                if v=="warning": return "color:#f59e0b"
                return "color:#60a5fa"
            st.dataframe(flag_df.style.map(sty_s,subset=["severity"]),
                         use_container_width=True, height=420)

    st.markdown("---")
    d1, d2, d3 = st.columns(3)

    if "phase3_enhanced" not in st.session_state or st.session_state.get("phase3_enhanced_key") != id(p3):
        raw_df_dl  = st.session_state["phase1_result"]["data_frame"]
        locked_dl  = st.session_state["locked_schema"]
        maps_dl    = st.session_state["phase1_result"]["mappings"]
        st.session_state["_cleaned_reason_bytes"] = build_cleaned_with_reasons(
            cleaned_df, raw_df_dl, locked_dl, flag_log,
            rules_df=p3.get("rules_df"))
        st.session_state["_enhanced_qa_bytes"]    = build_enhanced_qa_bytes(
            p3, locked_dl, maps_dl, raw_df_dl)
        st.session_state["phase3_enhanced"]     = True
        st.session_state["phase3_enhanced_key"] = id(p3)

    with d1:
        st.download_button("↓  Cleaned SOV with Reasons (.xlsx)",
                            data=st.session_state["_cleaned_reason_bytes"],
                            file_name=f"sov_cleaned_{system.lower()}_with_reasons.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True)
    with d2:
        st.download_button("↓  Cleaned SOV (.json)", data=p3["_json_str"],
                            file_name=f"sov_cleaned_{system.lower()}.json",
                            mime="application/json", use_container_width=True)
    with d3:
        st.download_button("↓  Full QA Report (.xlsx)",
                            data=st.session_state["_enhanced_qa_bytes"],
                            file_name=f"sov_qa_{system.lower()}_full.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True)


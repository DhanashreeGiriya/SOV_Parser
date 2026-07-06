"""
Auto-extracted module: header_mapping/reporting.py
"""

from __future__ import annotations

import openpyxl
import json
from pathlib import Path
from dataclasses import dataclass, field, asdict

from header_mapping.rms_crosswalk import AIR_TO_RMS_CONSTRUCTION, AIR_TO_RMS_OCCUPANCY, ISO_TO_AIR_CONSTRUCTION, RMS_COUNTRY_IND
from header_mapping.schema import AI_REVIEW_THRESHOLD

CONFIDENCE_COLOURS = {
    "high":    "C6EFCE",
    "medium":  "FFEB9C",
    "low":     "FFC7CE",
    "auto":    "DEEBF7",
    "none":    "F2F2F2",
    "feedback":"E8D5F5",   # soft purple for feedback matches
}


def _confidence_band(score: int, match_type: str) -> str:
    if match_type == "feedback_match":
        return "feedback"
    if score == 0:
        return "none"
    if score >= 85:
        return "high"
    if score >= 60:
        return "medium"
    return "low"


def export_mapping_report(mappings, unmapped_raw, flags, output_path,
                           sov_file_name="", target_system="AIR"):
    output_path = Path(output_path)
    wb_out = openpyxl.Workbook()
    sys_label = target_system.upper()
    h2_fill = openpyxl.styles.PatternFill("solid", fgColor="4472C4")
    h2_font = openpyxl.styles.Font(bold=True, color="FFFFFF")

    ws1 = wb_out.active
    ws1.title = "Header_Mapping"
    col_hdr = [
        f"Output Column ({sys_label})", "Source Column(s) in SOV",
        "How It Was Matched", "Confidence %", "Flag",
        "Semantic Suggestion", "AI Validated?", "AI Reasoning",
        "Decision Basis",
    ]
    ws1.append(col_hdr)
    for cell in ws1[1]:
        cell.fill = h2_fill
        cell.font = h2_font

    for m in mappings:
        band = _confidence_band(m.confidence, m.match_type)
        fill = openpyxl.styles.PatternFill("solid", fgColor=CONFIDENCE_COLOURS[band])
        ai_agreed = (
            "Validated" if m.ai_agreement else
            ("N/A" if not m.ai_suggestion else "Refined")
        )
        row = [
            m.output_col,
            " | ".join(m.source_cols) if m.source_cols else "—",
            m.match_type.replace("_", " ").title(),
            m.confidence,
            m.flag if m.flag else "OK",
            m.fuzzy_suggestion or m.alias_suggestion or "—",
            ai_agreed,
            m.ai_reasoning,
            m.final_decision_basis,
        ]
        ws1.append(row)
        for cell in ws1[ws1.max_row]:
            cell.fill = fill

    col_widths = [28, 40, 22, 12, 18, 28, 14, 50, 80]
    for idx, w in enumerate(col_widths, start=1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    ws_xw = wb_out.create_sheet("Code_Crosswalk")

    def _xw_header(ws, row_data):
        ws.append(row_data)
        for cell in ws[ws.max_row]:
            cell.fill = h2_fill
            cell.font = h2_font

    ws_xw.append(["=== CONSTRUCTION CODE CROSSWALK: ISO => AIR => RMS ==="])
    _xw_header(ws_xw, ["ISO Fire Code", "AIR ConstructionCode", "RMS ClassCode", "RMS Label"])
    for iso_code, air_code in sorted(ISO_TO_AIR_CONSTRUCTION.items()):
        rms = AIR_TO_RMS_CONSTRUCTION.get(air_code, {"rms_code": 1000, "rms_label": "Unknown"})
        ws_xw.append([iso_code, air_code, rms["rms_code"], rms["rms_label"]])

    ws_xw.append([])
    ws_xw.append(["=== AIR => RMS OCCUPANCY CROSSWALK ==="])
    _xw_header(ws_xw, ["AIR OccupancyCode", "RMS OccupancyType", "RMS Label"])
    for air_code, rms in sorted(AIR_TO_RMS_OCCUPANCY.items()):
        ws_xw.append([air_code, rms["rms_code"], rms["rms_label"]])

    ws_xw.append([])
    ws_xw.append(["=== RMS COUNTRY IND LOOKUP ==="])
    _xw_header(ws_xw, ["CountryISO", "Country Name", "4-Digit IND", "2-Digit IND", "Notes"])
    for iso, entry in sorted(RMS_COUNTRY_IND.items()):
        ws_xw.append([iso, entry["label"], entry["4digit"], entry["2digit"], entry["notes"]])

    for col_idx in range(1, 6):
        ws_xw.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 30

    ws2 = wb_out.create_sheet("Unmapped_Raw_Columns")
    ws2.append(["Raw SOV Column (Unmapped)", "Suggested Action"])
    for cell in ws2[1]:
        cell.fill = h2_fill
        cell.font = h2_font
    for col in unmapped_raw:
        ws2.append([col, "Review — may be supplementary data or can be discarded"])
    ws2.column_dimensions["A"].width = 45
    ws2.column_dimensions["B"].width = 55

    ws3 = wb_out.create_sheet("Summary_Flags")
    ws3.append(["Category", "Count", "Columns"])
    for cell in ws3[1]:
        cell.fill = h2_fill
        cell.font = h2_font
    for category, items in flags.items():
        ws3.append([category.replace("_", " ").title(), len(items), ", ".join(items)])
    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 80

    wb_out.save(output_path)

    json_path = output_path.with_suffix(".json")
    report_data = {
        "sov_file": sov_file_name,
        "target_system": sys_label,
        "total_output_columns": len(mappings),
        "ai_review_threshold": AI_REVIEW_THRESHOLD,
        "flags": flags,
        "mappings": [asdict(m) for m in mappings],
        "unmapped_raw_columns": unmapped_raw,
    }
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(report_data, f, indent=2, default=str)

    return output_path


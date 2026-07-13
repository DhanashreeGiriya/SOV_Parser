from __future__ import annotations
import pandas as pd
import openpyxl
from fuzzywuzzy import fuzz
import re
import string
from pathlib import Path
# import wordninja
from header_mapping.aliases import ALIAS_MAP
from header_mapping.schema import TARGET_SCHEMA_AIR, TARGET_SCHEMA_RMS

def _normalise(text: str) -> str:
    text = str(text).lower()
    text = text.translate(str.maketrans("", "", string.punctuation))
    return re.sub(r"\s+", " ", text).strip()
# def normalise_word

def detect_header_row(ws, max_scan_rows: int = 30, min_hit_ratio: float = 0.15) -> int:
    import math
    known_tokens: set[str] = (
        {_normalise(k) for k in ALIAS_MAP.keys()}
        | {_normalise(s["output_col"]) for s in TARGET_SCHEMA_AIR}
        | {_normalise(s["output_col"]) for s in TARGET_SCHEMA_RMS}
    )
    best_row, best_score = 1, 0.0
    for row_idx in range(1, max_scan_rows + 1):
        cells = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        non_empty = [str(v) for v in cells
                     if v not in (None, "", " ") and len(str(v).strip()) > 2]
        if len(non_empty) < 3:
            continue
        hits = 0
        for cell_val in non_empty:
            norm = _normalise(cell_val)
            if norm in known_tokens:
                hits += 1
                continue
            for known in known_tokens:
                if fuzz.token_set_ratio(norm, known) >= 70:
                    hits += 1
                    break
        ratio = hits / len(non_empty)
        combined = ratio * math.log2(1 + hits)
        if combined > best_score:
            best_score = combined
            best_row = row_idx
    return best_row if best_score > 0 else 1


def auto_detect_best_sheet(file_path):
    path = Path(file_path)
    try:
        xl = pd.ExcelFile(path)
        sheet_names = xl.sheet_names
    except Exception:
        return 0
    if len(sheet_names) == 1:
        return sheet_names[0]

    known_tokens: set[str] = (
        {_normalise(k) for k in ALIAS_MAP.keys()}
        | {_normalise(s["output_col"]) for s in TARGET_SCHEMA_AIR}
        | {_normalise(s["output_col"]) for s in TARGET_SCHEMA_RMS}
    )
    best_sheet = sheet_names[0]
    best_score = -1
    for name in sheet_names:
        try:
            df_raw = pd.read_excel(path, sheet_name=name, header=None, nrows=50, dtype=str)
            best_row_score = 0
            for row_idx in range(len(df_raw)):
                row_vals = [str(v) for v in df_raw.iloc[row_idx] if pd.notna(v) and str(v).strip()]
                if len(row_vals) < 3:
                    continue
                hits = sum(
                    1 for v in row_vals
                    if _normalise(v) in known_tokens
                    or any(fuzz.token_set_ratio(_normalise(v), t) >= 70 for t in known_tokens)
                )                # check logic
                row_score = hits / max(len(row_vals), 1)
                if row_score > best_row_score:
                    best_row_score = row_score
            if best_row_score > best_score:
                best_score = best_row_score
                best_sheet = name
        except Exception:
            continue
    return best_sheet


def _is_unnamed(col: str) -> bool:
    s = str(col).strip()
    return s.startswith("Unnamed:") or re.match(r"^Column\s*\d+$", s, re.I) is not None


def _has_mostly_unnamed(headers: list, threshold: float = 0.5) -> bool:
    if not headers:
        return False
    return sum(1 for h in headers if _is_unnamed(h)) / len(headers) > threshold


def _infer_header_from_values(df_noheader: pd.DataFrame) -> tuple[int, list[str]]:
    known_tokens: set[str] = (
        {_normalise(k) for k in ALIAS_MAP.keys()}
        | {_normalise(s["output_col"]) for s in TARGET_SCHEMA_AIR}
        | {_normalise(s["output_col"]) for s in TARGET_SCHEMA_RMS}
    )
    best_idx, best_score = 0, 0.0
    for row_idx in range(min(30, len(df_noheader))):
        row_vals = [str(v).strip() for v in df_noheader.iloc[row_idx] if pd.notna(v) and str(v).strip()]
        if len(row_vals) < 3:
            continue
        hits = 0
        for v in row_vals:
            norm = _normalise(v)
            if norm in known_tokens:
                hits += 1
                continue
            for t in known_tokens:
                if fuzz.token_set_ratio(norm, t) >= 70:
                    hits += 1
                    break
        ratio = hits / len(row_vals)
        if ratio > best_score:
            best_score = ratio
            best_idx = row_idx

    if best_score >= 0.15:
        col_names = [str(v).strip() if pd.notna(v) and str(v).strip() else f"_col_{i}"
                     for i, v in enumerate(df_noheader.iloc[best_idx])]
        return best_idx, col_names

    col_names = [f"_col_{i}" for i in range(len(df_noheader.columns))]
    return -1, col_names


def load_sov(file_path, sheet_name=None, max_scan_rows: int = 25):
    path = Path(file_path)
    if sheet_name is None:
        sheet_name = auto_detect_best_sheet(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if isinstance(sheet_name, int):
        ws = wb.worksheets[sheet_name]
    else:
        ws = wb[sheet_name]
    header_row = detect_header_row(ws, max_scan_rows=max_scan_rows)
    wb.close()

    df = pd.read_excel(
        path,
        sheet_name=sheet_name if isinstance(sheet_name, str) else sheet_name,
        header=header_row - 1,
        dtype=str,
    )
    df.dropna(how="all", inplace=True)
    df.reset_index(drop=True, inplace=True)
    raw_headers = list(df.columns)

    if _has_mostly_unnamed(raw_headers):
        df_raw = pd.read_excel(
            path,
            sheet_name=sheet_name if isinstance(sheet_name, str) else sheet_name,
            header=None,
            dtype=str,
        )
        detected_idx, new_col_names = _infer_header_from_values(df_raw)

        if detected_idx >= 0:
            header_row = detected_idx + 1
            df = pd.read_excel(
                path,
                sheet_name=sheet_name if isinstance(sheet_name, str) else sheet_name,
                header=detected_idx,
                dtype=str,
            )
            df.dropna(how="all", inplace=True)
            df.reset_index(drop=True, inplace=True)
            raw_headers = list(df.columns)
            raw_headers = [
                f"_col_{i}" if _is_unnamed(h) else h
                for i, h in enumerate(raw_headers)
            ]
            df.columns = raw_headers
        else:
            header_row = 1
            df = df_raw.copy()
            df.dropna(how="all", inplace=True)
            df.reset_index(drop=True, inplace=True)
            df.columns = [f"_col_{i}" for i in range(len(df.columns))]
            raw_headers = list(df.columns)
    else:
        raw_headers = [
            f"_col_{i}" if _is_unnamed(h) else h
            for i, h in enumerate(raw_headers)
        ]
        df.columns = raw_headers

    return df, header_row, raw_headers


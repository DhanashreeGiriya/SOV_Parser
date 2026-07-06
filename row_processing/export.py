"""
Auto-extracted module: row_processing/export.py
"""

from __future__ import annotations

import pandas as pd
import openpyxl
from pathlib import Path

from sov_app.row_processing.column_order import AIR_COLUMN_ORDER, RMS_COLUMN_ORDER
from sov_app.row_processing.flags import FlagLog, validate_cross_column_flags
from sov_app.row_processing.process_row import process_row

def run_value_transformation(
    phase1_result, locked_schema, target_system="AIR", output_dir=".",
    report_name="sov_cleaned_output", days_covered=365, default_country="US",
    lob_col="", apply_proration=False, proration_group_col=None,
):
    df_raw = phase1_result["data_frame"]
    df_raw_original = df_raw  # preserve true original for RAW columns in exports
    sys_label = target_system.upper()
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    flag_log = FlagLog()

    # ── Apply row-level feedback rules that must run BEFORE the code rule ────
    # (e.g. LocationName, Street) — these rewrite the raw source column(s)
    # so process_row's transform_*/resolve_* logic sees the corrected value.
    pre_rule_log = []
    try:
        from sov_app.feedback.row_feedback import apply_rules_to_raw as _apply_pre_rules
        df_raw, pre_rule_log = _apply_pre_rules(df_raw, locked_schema)
        if pre_rule_log:
            import logging as _log
            _log.getLogger(__name__).info(
                f"[sov_row_feedback] Pre-applied {len(pre_rule_log)} rule override(s) to raw data"
            )
    except ImportError:
        pass   # sov_row_feedback not installed — skip silently
    except Exception as _prfe:
        import logging as _log
        _log.getLogger(__name__).warning(f"[sov_row_feedback] apply_rules_to_raw failed: {_prfe}")

    cleaned_rows = []
    rules_rows   = []
    for idx, (_, row) in enumerate(df_raw.iterrows()):
        out_row, rules_row = process_row(
            row_idx=idx, row=row, schema=locked_schema, flag_log=flag_log,
            target_system=sys_label, days_covered=days_covered,
            default_country=default_country, lob_col=lob_col,
        )
        cleaned_rows.append(out_row)
        rules_rows.append(rules_row)

    cleaned_df = pd.DataFrame(cleaned_rows)
    rules_df   = pd.DataFrame(rules_rows)
    validate_cross_column_flags(cleaned_df, flag_log)

    # ── Apply row-level feedback rules (post-processing, for all OTHER columns) ──
    rule_application_log = list(pre_rule_log)
    try:
        from sov_app.feedback.row_feedback import apply_rules as _apply_row_rules
        cleaned_df, post_rule_log = _apply_row_rules(
            cleaned_df, locked_schema, df_raw
        )
        rule_application_log += post_rule_log
        if post_rule_log:
            import logging as _log
            _log.getLogger(__name__).info(
                f"[sov_row_feedback] Applied {len(post_rule_log)} rule override(s)"
            )
    except ImportError:
        pass   # sov_row_feedback not installed — skip silently
    except Exception as _rfe:
        import logging as _log
        _log.getLogger(__name__).warning(f"[sov_row_feedback] apply_rules failed: {_rfe}")

    # Cast year / integer columns to nullable Int64 so NaN rows don't
    # force the column to float64 (which would show 2014.0 instead of 2014).
    _INT_COLS = [
        "YearBuilt", "Roof Year Built", "RoofCoverYear",
        "NumberOfStories", "NumStories",
        "RiskCount", "NumUnits", "DaysCovered",
        "Sprinkler Availability", "SprinklerType",
    ]
    for _col in _INT_COLS:
        if _col in cleaned_df.columns:
            cleaned_df[_col] = pd.to_numeric(cleaned_df[_col], errors="coerce")                                   .astype("Int64")
    
    
    
    # ── Preserve occupancy review helper columns before reindexing ──────────
    _OCC_HELPER_COLS = ["_occ_needs_review", "_occ_method", "_occ_confidence", "_occ_raw_description",
                     "_constr_needs_review", "_constr_method", "_constr_confidence", "_constr_raw_description"]
    _occ_helpers_df = cleaned_df[[c for c in _OCC_HELPER_COLS if c in cleaned_df.columns]].copy()

    col_order = AIR_COLUMN_ORDER if sys_label == "AIR" else RMS_COLUMN_ORDER
    for col in col_order:
        if col not in cleaned_df.columns:
            cleaned_df[col] = None
    cleaned_df = cleaned_df[[c for c in col_order if c in cleaned_df.columns]]

    # ── Re-attach occupancy helper columns so the review panel can read them ──
    for c in _OCC_HELPER_COLS:
        if c in _occ_helpers_df.columns:
            cleaned_df[c] = _occ_helpers_df[c].values

    excel_path = output_dir / f"{report_name}_{sys_label.lower()}_cleaned.xlsx"
    json_path  = output_dir / f"{report_name}_{sys_label.lower()}_cleaned.json"
    qa_path    = output_dir / f"{report_name}_{sys_label.lower()}_qa_report.xlsx"

    _export_cleaned_excel(cleaned_df, excel_path, flag_log, sys_label,
                          raw_df=df_raw_original, locked_schema=locked_schema)
    _EXPORT_EXCLUDE = {"_occ_needs_review", "_occ_method", "_occ_confidence", "_occ_raw_description",
                    "_constr_needs_review", "_constr_method", "_constr_confidence", "_constr_raw_description"}
    _export_df = cleaned_df[[c for c in cleaned_df.columns if c not in _EXPORT_EXCLUDE]]
    _export_df.to_json(json_path, orient="records", indent=2, default_handler=str)
    _export_qa_report(cleaned_df, flag_log, qa_path, sys_label)

    return {
        "cleaned_df":          cleaned_df,
        "rules_df":            rules_df,
        "flag_log":            flag_log,
        "flag_summary":        flag_log.summary(),
        "output_excel":        str(excel_path),
        "output_json":         str(json_path),
        "qa_report":           str(qa_path),
        "rule_application_log":rule_application_log,
    }


def _export_cleaned_excel(df, path, flag_log, sys_label, raw_df=None, locked_schema=None):
    _EXPORT_EXCLUDE = {"_occ_needs_review", "_occ_method", "_occ_confidence", "_occ_raw_description"}
    df = df[[c for c in df.columns if c not in _EXPORT_EXCLUDE]]
    SEV_COLOUR = {"error": "FFC7CE", "warning": "FFEB9C", "info": "DEEBF7"}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Cleaned_{sys_label}"

    out_to_sources = {}
    if locked_schema is not None:
        for d in locked_schema.decisions:
            out_to_sources[d.output_col] = d.final_source or []

    interleaved = []
    for out_col in df.columns:
        sources = out_to_sources.get(out_col, [])
        for src_col in sources:
            if raw_df is not None and src_col in raw_df.columns:
                interleaved.append((f"RAW: {src_col}", src_col, True))
        interleaved.append((out_col, out_col, False))

    RAW_HDR = "D9E1F2"
    OUT_HDR = "4472C4"
    for col_idx, (label, _, is_raw) in enumerate(interleaved, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        colour = RAW_HDR if is_raw else OUT_HDR
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor=colour)
        cell.font = openpyxl.styles.Font(
            bold=True, color="333333" if is_raw else "FFFFFF", italic=is_raw)

    flag_lookup = {}
    for f in flag_log._flags:
        key = (f.row_idx, f.output_col)
        existing = flag_lookup.get(key, "info")
        priority = {"error": 2, "warning": 1, "info": 0}
        if priority.get(f.severity, 0) > priority.get(existing, 0):
            flag_lookup[key] = f.severity

    for row_idx, (_, out_row) in enumerate(df.iterrows()):
        raw_row = raw_df.iloc[row_idx] if raw_df is not None else None
        excel_row = row_idx + 2
        for col_idx, (label, col_name, is_raw) in enumerate(interleaved, 1):
            if is_raw:
                val = raw_row.get(col_name, "") if raw_row is not None else ""
                cell = ws.cell(excel_row, col_idx, value=str(val) if pd.notna(val) else "")
                cell.fill = openpyxl.styles.PatternFill("solid", fgColor="F2F7FF")
            else:
                val = out_row.get(col_name)
                # Convert pandas NA / numpy int64 to plain Python types so
                # openpyxl doesn't write "2014.0" or raise on pd.NA.
                import pandas as _pd
                if _pd.isna(val) if not isinstance(val, (list, dict)) else False:
                    val = None
                elif hasattr(val, "item"):          # numpy scalar -> Python scalar
                    val = val.item()
                cell = ws.cell(excel_row, col_idx, value=val)
                sev = flag_lookup.get((row_idx, col_name))
                if sev:
                    cell.fill = openpyxl.styles.PatternFill(
                        "solid", fgColor=SEV_COLOUR.get(sev, "F2F2F2"))

    for i, (label, _, _) in enumerate(interleaved, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(len(str(label)) + 2, 14)
    ws.freeze_panes = "A2"
    wb.save(path)


def _export_qa_report(df, flag_log, path, sys_label):
    wb = openpyxl.Workbook()
    hdr_fill = openpyxl.styles.PatternFill("solid", fgColor="4472C4")
    hdr_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    err_fill  = openpyxl.styles.PatternFill("solid", fgColor="FFC7CE")
    warn_fill = openpyxl.styles.PatternFill("solid", fgColor="FFEB9C")
    info_fill = openpyxl.styles.PatternFill("solid", fgColor="DEEBF7")

    flag_df = flag_log.to_dataframe()
    total_rows = len(df)
    total_flags = len(flag_df)
    flagged_rows = flag_df["row_idx"].nunique() if not flag_df.empty else 0
    clean_rows = total_rows - flagged_rows

    def write_header(ws, headers):
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.fill = hdr_fill
            cell.font = hdr_font

    def autowidth(ws):
        for col_cells in ws.columns:
            width = 12
            for cell in col_cells:
                try:
                    width = max(width, min(len(str(cell.value or "")), 60))
                except Exception:
                    pass
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_cells[0].column)].width = width + 2

    ws1 = wb.active
    ws1.title = "Summary"
    write_header(ws1, ["Metric", "Value", "Status"])
    for row in [
        ("Total Locations", total_rows, ""),
        ("Clean Rows", clean_rows, "OK" if clean_rows == total_rows else ""),
        ("Flagged Rows", flagged_rows, "Review" if flagged_rows > 0 else "OK"),
        ("Total Flags", total_flags, ""),
        ("Errors", len(flag_df[flag_df.severity == "error"]) if not flag_df.empty else 0, ""),
    ]:
        ws1.append(list(row))
    autowidth(ws1)

    ws2 = wb.create_sheet("Flag_Log")
    if not flag_df.empty:
        write_header(ws2, list(flag_df.columns))
        for _, r in flag_df.iterrows():
            ws2.append(list(r))
            sev = r.get("severity", "")
            fill = err_fill if sev == "error" else (warn_fill if sev == "warning" else info_fill)
            for cell in ws2[ws2.max_row]:
                cell.fill = fill
    else:
        ws2.append(["No flags raised."])
    autowidth(ws2)

    wb.save(path)

